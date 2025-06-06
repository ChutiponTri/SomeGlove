from PyQt6.QtCore import pyqtSignal, QThread
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
from datetime import datetime
import numpy as np
import time
import math
import os
from demo_database import Database

class Analysis(QThread):
    # Create pyqtSignal
    final_data = pyqtSignal(list,list,float,float,float,float,float,float,int)
    hr_data = pyqtSignal(float,float,float,int)
    save_status = pyqtSignal(str)

    # Initialize Function
    def __init__(self, data_to_save):
        super().__init__()
        self.username = data_to_save["Username"]
        self.pressure = data_to_save["Pressure"]
        self.usage_time = (datetime.fromisoformat(data_to_save["Stop Time"]) - datetime.fromisoformat(data_to_save["Start Time"])).total_seconds()
        self.start_time = data_to_save["Start Time"]
        self.stop_time = data_to_save["Stop Time"]
        self.game_label = data_to_save["Game"]

    # Auto Call Function When The start() method Called
    def run(self):
        self.grab = self.calculate_grab()
        self.pressure = [self.analog_to_pressure(i) for i in self.pressure]
        flow_rate = self.calculate_flow_rate(0.7, 25.97)
        self.power = self.calculate_power(np.mean(np.abs(self.pressure)), flow_rate)
        self.energy = self.calculate_energy(self.power, self.usage_time)
        self.calorie = self.calculate_calories(self.energy)
        self.save()

    # Function to Count Grabs
    def calculate_grab(self):
        threshold_drop = 200  # Minimum drop to consider a valley
        window_size = 5       # Size of the window to check

        valleys = 0
        i = 0

        while i < len(self.pressure) - window_size:
            found_valley = False
            for j in range(i + 1, i + window_size):
                if self.pressure[j] - self.pressure[i] < - threshold_drop:
                    valleys += 1
                    found_valley = True
                    break  
            if found_valley:
                i += window_size
            else:
                i += 1
        return valleys
    
    # Function to calculate Pressure
    def analog_to_pressure(self, analog_value, v_ref=3.3, offset=0.0):
        min_kPa = -40  # ค่า kPa ต่ำสุด
        max_kPa = 40   # ค่า kPa สูงสุด
    
        pressure_range = max_kPa - min_kPa
        voltage = (analog_value / 4095) * v_ref
        kPa = ((analog_value / 4095) * pressure_range + min_kPa) + offset
        mmHg = kPa * 7.50062
    
        return kPa
    
    # Function to Calculate Flow Rate
    def calculate_flow_rate(self, diameter_cm, velocity_m_s):
        diameter_m = diameter_cm / 100  # Convert diameter to meters
        area_m2 = math.pi * (diameter_m / 2) ** 2
        flow_rate = area_m2 * velocity_m_s
        return flow_rate

    # Function to Calculate Power in Watts
    def calculate_power(self, pressure_kPa, flow_rate):
        pressure_Pa = pressure_kPa * 1000  # Convert kPa to Pa
        power_W = pressure_Pa * flow_rate
        return power_W
    
    # Function to Calculate Energy in Joules
    def calculate_energy(self, power_W, time_s):
        energy_J = power_W * time_s
        return energy_J

    # Function to Calculate Calories
    def calculate_calories(self, energy_J):
        kcal = energy_J / 4184  # 1 kcal = 4184 J
        return kcal
    
    # Function to Calculate Power
    def calculate_power(self, pressure, flow_rate):
        pressure_Pa = pressure * 1000  # Convert kPa to Pa
        power_W = pressure_Pa * flow_rate
        return power_W
    
    # Function to save
    def save(self):
        try:
            Database.insert_overview(self.start_time, self.stop_time, self.username, self.usage_time, self.grab, np.mean(self.pressure), np.max(self.pressure), self.power, self.calorie, self.game_label)
            # file_name = "%s_overview.xlsx" % self.username
            # file_path = "excel/%s_overview.xlsx" % self.username
            # data_path = "excel/%s-%s-%s.xlsx" % (self.username, self.game_label, self.start_time)
            # folder = os.listdir("excel")
            # overall_dict = {
            #         "เวลาขณะเริ่มต้น": [self.start_time],
            #         "เวลาขณะเสร็จสิ้น": [self.stop_time],
            #         "เวลาที่ใช้ (นาที)": [self.usage_time],
            #         "จำนวนครั้งที่กำมือ": [self.grab],
            #         "ความดันเฉลี่ย (kPa)": [np.mean(self.pressure)],
            #         "ความดันสูงสุด (kPa)": [np.max(self.pressure)],
            #         "พลังงาน (W)": [self.power],
            #         "แคลอรี่ (kcal)": [self.calorie],
            #         "เกมที่ทดสอบ": [self.game_label]
            #     }
            # new_df = pd.DataFrame(overall_dict)
            # if file_name in folder:
            #     print("%s excel file is exist" % file_name)
            #     exist_df = pd.read_excel(file_path)
            #     combine_df = pd.concat([exist_df, new_df]).reset_index(drop=True)
            #     combine_df.to_excel(file_path, index=False)

            # else:
            #     print("%s excel file is not exist" % file_name)
            #     new_df.to_excel(file_path, index=False)

            # data_dict = {"ความดัน": self.pressure}
            # data_df = pd.DataFrame(data_dict)
            # data_df.to_excel(data_path, index=False)

            self.save_status.emit("Save")
            print("Saved %s", self.username)
            # print("Save file %s" % file_name)
                
        except PermissionError:
            print("Please Close Excel File Before Saving")
            self.save_status.emit("Permission")
            time.sleep(2)

        except Exception as e:
            print(print(f"An unexpected error occurred: {e}"))
            self.save_status.emit("Error %s" % e)

    # Function to save
    def old_save(self):
        saving_success = False
        while not saving_success:
            try:
                file_name = "%s.xlsx" % self.username
                file_path = "excel/%s.xlsx" % self.username
                folder = os.listdir("excel")
                if file_name in folder:
                    print("%s excel file is exist" % file_name)
                    workbook = load_workbook(file_path)
                    overview = workbook["ภาพรวม"]
                    index = str(len(overview["A"]) + 1)

                    overview["A" + index] = self.start_time
                    overview["B" + index] = self.stop_time
                    overview["C" + index] = self.usage_time
                    overview["D" + index] = self.grab
                    overview["E" + index] = np.mean(self.pressure)
                    overview["F" + index] = np.max(self.pressure)
                    overview["G" + index] = self.power
                    overview["H" + index] = self.calorie
                    overview["I" + index] = self.game_label

                    title = "data_%s" % self.start_time
                    workbook.create_sheet(title=title)

                    # Append Column To Sheet
                    sheet = workbook[title]
                    sheet.append(["ความดัน"])
                    header = sheet[1]
                    for cell in header:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Append Data To Sheet
                    for row in self.pressure:
                        sheet.append([row])

                    # Adjust Sheet Cell Width
                    sheet.column_dimensions["A"].width = 20

                    # Save Excel
                    workbook.save(file_path)
                    saving_success = True

                    self.save_status.emit("Save")
                    
            except PermissionError:
                print("Please Close Excel File Before Saving")
                self.save_status.emit("Permission")
                time.sleep(2)

            except Exception as e:
                print(print(f"An unexpected error occurred: {e}"))
                self.save_status.emit("Error %s" % e)

        print("Save file %s" % file_name)